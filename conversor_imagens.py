#!/usr/bin/env python3
"""
Conversor de imagens com interface PyQt5.
Localiza imagens em uma pasta, converte para o formato escolhido
e salva uma cópia no mesmo diretório do arquivo original.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import Cell
from PIL import Image
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# Extensões suportadas (leitura)
EXTENSOES_IMAGEM = {
    ".png",
    ".jpg",
    ".jpeg",
    ".jpe",
    ".bmp",
    ".gif",
    ".tif",
    ".tiff",
    ".webp",
    ".ico",
    ".icns",
    ".ppm",
    ".pgm",
    ".pbm",
    ".tga",
    ".dds",
    ".dib",
}

# Formatos de saída disponíveis na interface
FORMATOS_SAIDA = [
    ("PNG", "png", "PNG"),
    ("JPEG", "jpg", "JPEG"),
    ("WEBP", "webp", "WEBP"),
    ("BMP", "bmp", "BMP"),
    ("TIFF", "tiff", "TIFF"),
    ("GIF", "gif", "GIF"),
    ("ICO", "ico", "ICO"),
]

# Filtro de tipo de origem (o que buscar/converter)
# None = todos os formatos de imagem suportados
FILTROS_ORIGEM = [
    ("Todos os formatos", None),
    ("Somente ICO", {".ico"}),
    ("Somente PNG", {".png"}),
    ("Somente JPEG", {".jpg", ".jpeg", ".jpe"}),
    ("Somente WEBP", {".webp"}),
    ("Somente BMP", {".bmp", ".dib"}),
    ("Somente GIF", {".gif"}),
    ("Somente TIFF", {".tif", ".tiff"}),
]


def listar_imagens(pasta: Path, recursivo: bool) -> list[Path]:
    """Encontra todos os arquivos de imagem na pasta."""
    encontrados: list[Path] = []
    if recursivo:
        candidatos = pasta.rglob("*")
    else:
        candidatos = pasta.iterdir()

    for caminho in candidatos:
        if caminho.is_file() and caminho.suffix.lower() in EXTENSOES_IMAGEM:
            encontrados.append(caminho)

    return sorted(encontrados, key=lambda p: str(p).lower())


def preparar_imagem(img: Image.Image, formato_pil: str) -> Image.Image:
    """Ajusta modo de cor conforme o formato de destino."""
    formato = formato_pil.upper()

    # ICO costuma trazer várias resoluções; usamos o maior frame
    if getattr(img, "n_frames", 1) > 1 and formato != "GIF":
        melhor = img
        melhor_area = img.size[0] * img.size[1]
        for i in range(img.n_frames):
            img.seek(i)
            area = img.size[0] * img.size[1]
            if area > melhor_area:
                melhor = img.copy()
                melhor_area = area
        img = melhor if melhor is not img else img.copy()

    if formato in {"JPEG", "JPG"}:
        if img.mode in ("RGBA", "LA", "P"):
            fundo = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            alpha = img.split()[-1] if img.mode in ("RGBA", "LA") else None
            fundo.paste(img, mask=alpha)
            return fundo
        if img.mode != "RGB":
            return img.convert("RGB")
        return img

    if formato == "BMP" and img.mode in ("RGBA", "LA", "P"):
        return img.convert("RGB")

    if formato == "ICO":
        # ICO com transparência funciona melhor em RGBA
        if img.mode != "RGBA":
            return img.convert("RGBA")
        return img

    return img


def converter_imagem(origem: Path, extensao: str, formato_pil: str) -> Path:
    """
    Converte a imagem e salva uma cópia ao lado do original.
    Ex.: foto.ico -> foto.png (mantém o .ico intacto).
    """
    destino = origem.with_suffix(f".{extensao}")

    # Evita sobrescrever se já existir com outro conteúdo
    if destino.resolve() == origem.resolve():
        destino = origem.with_name(f"{origem.stem}_convertido.{extensao}")

    with Image.open(origem) as img:
        img.load()
        processada = preparar_imagem(img, formato_pil)

        opcoes: dict = {}
        if formato_pil.upper() in {"JPEG", "JPG"}:
            opcoes["quality"] = 95
            opcoes["optimize"] = True
        elif formato_pil.upper() == "WEBP":
            opcoes["quality"] = 90
        elif formato_pil.upper() == "ICO":
            # Gera ícone com tamanhos comuns
            opcoes["sizes"] = [(16, 16), (32, 32), (48, 48), (64, 64), (128, 128), (256, 256)]

        processada.save(destino, format=formato_pil, **opcoes)

    return destino


def _uri_local(caminho: Path | str) -> str:
    """Converte caminho Windows em URI file:// para hiperlink no Excel."""
    return Path(caminho).resolve().as_uri()


def _aplicar_link(celula: Cell, caminho: Path | str, texto: str | None = None) -> None:
    """Coloca texto clicável que abre o arquivo/pasta no Explorer."""
    caminho_p = Path(caminho)
    celula.value = texto if texto is not None else str(caminho_p.resolve())
    celula.hyperlink = _uri_local(caminho_p)
    celula.font = Font(color="0563C1", underline="single")


def exportar_excel(
    arquivos: list[Path],
    destino_xlsx: Path,
    pasta_origem: Path | None = None,
    gerados: list[str] | None = None,
) -> Path:
    """Gera um Excel com caminhos clicáveis (hiperlinks) das imagens."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Imagens"

    cabecalhos = [
        "Nº",
        "Arquivo",
        "Formato",
        "Pasta",
        "Caminho completo",
        "Cópia convertida",
    ]
    ws.append(cabecalhos)
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(1, col).font = Font(bold=True)

    mapa_gerados = {Path(g).stem.lower(): g for g in (gerados or [])}

    for i, caminho in enumerate(arquivos, start=1):
        linha = i + 1
        caminho_abs = caminho.resolve()
        pasta = caminho_abs.parent
        copia = mapa_gerados.get(caminho.stem.lower(), "")

        ws.cell(linha, 1, i)
        _aplicar_link(ws.cell(linha, 2), caminho_abs, caminho.name)
        ws.cell(linha, 3, caminho.suffix.lower().lstrip("."))
        _aplicar_link(ws.cell(linha, 4), pasta, str(pasta))
        _aplicar_link(ws.cell(linha, 5), caminho_abs, str(caminho_abs))

        celula_copia = ws.cell(linha, 6)
        if copia:
            copia_p = Path(copia)
            if copia_p.exists():
                _aplicar_link(celula_copia, copia_p, str(copia_p.resolve()))
            else:
                celula_copia.value = str(copia_p)
        else:
            celula_copia.value = ""

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 80

    if pasta_origem is not None:
        info = wb.create_sheet("Info")
        info["A1"] = "Pasta de origem"
        info["A1"].font = Font(bold=True)
        _aplicar_link(info["B1"], pasta_origem.resolve())
        info["A2"] = "Total de imagens"
        info["A2"].font = Font(bold=True)
        info["B2"] = len(arquivos)
        info["A4"] = "Dica"
        info["B4"] = (
            "Clique nos links azuis (Arquivo, Pasta, Caminho) para abrir no Explorer."
        )
        info.column_dimensions["A"].width = 22
        info.column_dimensions["B"].width = 90

    destino_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino_xlsx)
    return destino_xlsx


class WorkerConversao(QThread):
    progresso = pyqtSignal(int, int, str)
    finalizado = pyqtSignal(int, int, list)
    erro_arquivo = pyqtSignal(str, str)

    def __init__(self, arquivos: list[Path], extensao: str, formato_pil: str):
        super().__init__()
        self.arquivos = arquivos
        self.extensao = extensao
        self.formato_pil = formato_pil
        self._cancelar = False

    def cancelar(self) -> None:
        self._cancelar = True

    def run(self) -> None:
        ok = 0
        falhas = 0
        gerados: list[str] = []
        total = len(self.arquivos)

        for i, arquivo in enumerate(self.arquivos, start=1):
            if self._cancelar:
                break
            self.progresso.emit(i, total, str(arquivo))
            try:
                destino = converter_imagem(arquivo, self.extensao, self.formato_pil)
                gerados.append(str(destino))
                ok += 1
            except Exception as exc:  # noqa: BLE001 — reporta erro por arquivo
                falhas += 1
                self.erro_arquivo.emit(str(arquivo), str(exc))

        self.finalizado.emit(ok, falhas, gerados)


class JanelaConversor(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("ConversorImagens")
        self.resize(920, 640)
        self.arquivos: list[Path] = []
        self.ultimos_gerados: list[str] = []
        self.worker: WorkerConversao | None = None

        raiz = QWidget()
        self.setCentralWidget(raiz)
        layout = QVBoxLayout(raiz)
        layout.setSpacing(10)

        # --- Pasta ---
        grupo_pasta = QGroupBox("Pasta de origem")
        layout_pasta = QHBoxLayout(grupo_pasta)
        self.campo_pasta = QLineEdit()
        self.campo_pasta.setPlaceholderText("Selecione a pasta com as imagens...")
        self.campo_pasta.setReadOnly(True)
        btn_pasta = QPushButton("Procurar...")
        btn_pasta.clicked.connect(self.escolher_pasta)
        layout_pasta.addWidget(self.campo_pasta)
        layout_pasta.addWidget(btn_pasta)
        layout.addWidget(grupo_pasta)

        # --- Opções ---
        grupo_opcoes = QGroupBox("Opções de conversão")
        layout_opcoes = QHBoxLayout(grupo_opcoes)

        layout_opcoes.addWidget(QLabel("Converter de:"))
        self.combo_origem = QComboBox()
        for rotulo, extensoes in FILTROS_ORIGEM:
            self.combo_origem.addItem(rotulo, extensoes)
        self.combo_origem.setToolTip(
            "Escolha o tipo de arquivo que deseja buscar/converter "
            "(a pasta pode ter PNG, JPEG, ICO e outros misturados)."
        )
        # Padrão: Somente ICO (índice 1)
        self.combo_origem.setCurrentIndex(1)
        layout_opcoes.addWidget(self.combo_origem)

        layout_opcoes.addWidget(QLabel("Para:"))
        self.combo_formato = QComboBox()
        for rotulo, extensao, formato_pil in FORMATOS_SAIDA:
            self.combo_formato.addItem(rotulo, (extensao, formato_pil))
        # PNG como padrão (útil para ICO → PNG)
        self.combo_formato.setCurrentIndex(0)
        layout_opcoes.addWidget(self.combo_formato)

        self.check_recursivo = QCheckBox("Incluir subpastas")
        self.check_recursivo.setChecked(True)
        layout_opcoes.addWidget(self.check_recursivo)

        layout_opcoes.addStretch()
        layout.addWidget(grupo_opcoes)

        # --- Botões de ação ---
        layout_botoes = QHBoxLayout()
        self.btn_buscar = QPushButton("1. Buscar imagens")
        self.btn_buscar.clicked.connect(self.buscar_imagens)
        self.btn_converter = QPushButton("2. Converter e salvar cópias")
        self.btn_converter.clicked.connect(self.iniciar_conversao)
        self.btn_converter.setEnabled(False)
        self.btn_excel = QPushButton("3. Salvar Excel (caminhos)")
        self.btn_excel.setToolTip("Salva um .xlsx com o caminho de todas as imagens listadas")
        self.btn_excel.clicked.connect(self.salvar_excel)
        self.btn_excel.setEnabled(False)
        self.btn_cancelar = QPushButton("Cancelar")
        self.btn_cancelar.clicked.connect(self.cancelar_conversao)
        self.btn_cancelar.setEnabled(False)
        layout_botoes.addWidget(self.btn_buscar)
        layout_botoes.addWidget(self.btn_converter)
        layout_botoes.addWidget(self.btn_excel)
        layout_botoes.addWidget(self.btn_cancelar)
        layout.addLayout(layout_botoes)

        # --- Tabela ---
        self.label_resumo = QLabel("Nenhuma imagem listada.")
        layout.addWidget(self.label_resumo)

        self.tabela = QTableWidget(0, 4)
        self.tabela.setHorizontalHeaderLabels(["Arquivo", "Formato", "Pasta", "Caminho completo"])
        self.tabela.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabela.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tabela.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.tabela.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tabela.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabela.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.tabela)

        # --- Progresso e log ---
        self.barra = QProgressBar()
        self.barra.setValue(0)
        layout.addWidget(self.barra)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumHeight(140)
        self.log.setPlaceholderText("Log de conversão...")
        layout.addWidget(self.log)

        dica = QLabel(
            "A pasta pode ter PNG, JPEG, ICO e outros juntos. "
            "Em 'Converter de' escolha o tipo que quer processar; "
            "em 'Para' escolha o formato de saída. "
            "A cópia fica no mesmo local (ex.: icone.ico → icone.png)."
        )
        dica.setWordWrap(True)
        dica.setStyleSheet("color: #555;")
        layout.addWidget(dica)

    def escolher_pasta(self) -> None:
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar pasta com imagens")
        if pasta:
            self.campo_pasta.setText(pasta)
            self.buscar_imagens()

    def buscar_imagens(self) -> None:
        pasta_txt = self.campo_pasta.text().strip()
        if not pasta_txt:
            QMessageBox.warning(self, "Atenção", "Selecione uma pasta primeiro.")
            return

        pasta = Path(pasta_txt)
        if not pasta.is_dir():
            QMessageBox.critical(self, "Erro", "A pasta selecionada não existe.")
            return

        arquivos = listar_imagens(pasta, self.check_recursivo.isChecked())
        filtro_ext = self.combo_origem.currentData()
        if filtro_ext:
            arquivos = [a for a in arquivos if a.suffix.lower() in filtro_ext]

        self.arquivos = arquivos
        self.tabela.setRowCount(0)

        for caminho in arquivos:
            row = self.tabela.rowCount()
            self.tabela.insertRow(row)
            self.tabela.setItem(row, 0, QTableWidgetItem(caminho.name))
            self.tabela.setItem(row, 1, QTableWidgetItem(caminho.suffix.lower().lstrip(".")))
            self.tabela.setItem(row, 2, QTableWidgetItem(str(caminho.parent)))
            self.tabela.setItem(row, 3, QTableWidgetItem(str(caminho)))

        filtro_txt = self.combo_origem.currentText()
        self.ultimos_gerados = []
        self.label_resumo.setText(
            f"{len(arquivos)} imagem(ns) encontrada(s) — filtro: {filtro_txt}."
        )
        self.btn_converter.setEnabled(len(arquivos) > 0)
        self.btn_excel.setEnabled(len(arquivos) > 0)
        self.barra.setValue(0)
        self.log.append(
            f"Busca concluída: {len(arquivos)} arquivo(s) em {pasta} [{filtro_txt}]"
        )

        if not arquivos:
            QMessageBox.information(
                self,
                "Nada encontrado",
                "Nenhuma imagem foi encontrada com os filtros atuais.",
            )

    def iniciar_conversao(self) -> None:
        if not self.arquivos:
            return

        extensao, formato_pil = self.combo_formato.currentData()
        resposta = QMessageBox.question(
            self,
            "Confirmar conversão",
            (
                f"Converter {len(self.arquivos)} imagem(ns) para .{extensao.upper()}?\n\n"
                "As cópias serão salvas no mesmo local das originais."
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if resposta != QMessageBox.Yes:
            return

        self.btn_buscar.setEnabled(False)
        self.btn_converter.setEnabled(False)
        self.btn_excel.setEnabled(False)
        self.btn_cancelar.setEnabled(True)
        self.barra.setMaximum(len(self.arquivos))
        self.barra.setValue(0)
        self.log.append(f"Iniciando conversão para {formato_pil}...")

        self.worker = WorkerConversao(self.arquivos, extensao, formato_pil)
        self.worker.progresso.connect(self._on_progresso)
        self.worker.erro_arquivo.connect(self._on_erro)
        self.worker.finalizado.connect(self._on_finalizado)
        self.worker.start()

    def salvar_excel(self) -> None:
        if not self.arquivos:
            QMessageBox.warning(self, "Atenção", "Busque as imagens antes de salvar o Excel.")
            return

        pasta_txt = self.campo_pasta.text().strip()
        sugestao = Path(pasta_txt or ".") / "lista_imagens.xlsx"
        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Excel com caminhos das imagens",
            str(sugestao),
            "Excel (*.xlsx)",
        )
        if not caminho:
            return

        destino = Path(caminho)
        if destino.suffix.lower() != ".xlsx":
            destino = destino.with_suffix(".xlsx")

        try:
            pasta_origem = Path(pasta_txt) if pasta_txt else None
            exportar_excel(
                self.arquivos,
                destino,
                pasta_origem=pasta_origem,
                gerados=self.ultimos_gerados,
            )
            self.log.append(f"Excel salvo: {destino}")
            QMessageBox.information(
                self,
                "Excel salvo",
                f"Arquivo gerado com {len(self.arquivos)} caminho(s):\n\n{destino}",
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "Erro ao salvar Excel", str(exc))

    def cancelar_conversao(self) -> None:
        if self.worker and self.worker.isRunning():
            self.worker.cancelar()
            self.log.append("Cancelamento solicitado...")

    def _on_progresso(self, atual: int, total: int, caminho: str) -> None:
        self.barra.setMaximum(total)
        self.barra.setValue(atual)
        self.statusBar().showMessage(f"Convertendo ({atual}/{total}): {caminho}")

    def _on_erro(self, caminho: str, mensagem: str) -> None:
        self.log.append(f"ERRO: {caminho} → {mensagem}")

    def _on_finalizado(self, ok: int, falhas: int, gerados: list) -> None:
        self.ultimos_gerados = list(gerados)
        self.btn_buscar.setEnabled(True)
        self.btn_converter.setEnabled(bool(self.arquivos))
        self.btn_excel.setEnabled(bool(self.arquivos))
        self.btn_cancelar.setEnabled(False)
        self.statusBar().showMessage("Conversão finalizada.", 5000)

        for caminho in gerados:
            self.log.append(f"OK: {caminho}")

        self.log.append(f"Resumo: {ok} convertida(s), {falhas} falha(s).")

        msg = f"Conversão finalizada.\n\nSucesso: {ok}\nFalhas: {falhas}"
        if ok > 0 and self.arquivos:
            perguntar = QMessageBox.question(
                self,
                "Concluído",
                msg + "\n\nDeseja salvar o Excel com os caminhos agora?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if perguntar == QMessageBox.Yes:
                self.salvar_excel()
        else:
            QMessageBox.information(self, "Concluído", msg)


def main() -> None:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    janela = JanelaConversor()
    janela.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
